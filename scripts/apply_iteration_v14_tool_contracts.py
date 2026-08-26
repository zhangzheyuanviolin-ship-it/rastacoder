#!/usr/bin/env python3
"""Apply V14 model-facing creation contracts and semantic XLSX hardening."""
from pathlib import Path


def replace_once(text: str, old: str, new: str, label: str) -> str:
    count = text.count(old)
    if count != 1:
        raise SystemExit(f"{label}: expected exactly one anchor, found {count}")
    return text.replace(old, new, 1)


compat_path = Path("python/navixmind/tools/compat.py")
compat = compat_path.read_text()
if "RASTACODER_V14_STRUCTURED_CREATION_COMPAT" not in compat:
    anchor = '''    if name == "write_file":
        _move_alias(args, "content", ["text", "body", "data"], notes)
'''
    block = '''    # RASTACODER_V14_STRUCTURED_CREATION_COMPAT
    if name == "create_xlsx":
        sheets = args.get("sheets")
        if isinstance(sheets, dict):
            sheets = [sheets]
            notes.append("sheets:object->list")
        if isinstance(sheets, (list, tuple)):
            normalized_sheets = []
            for index, raw_spec in enumerate(sheets):
                if not isinstance(raw_spec, dict):
                    normalized_sheets.append(raw_spec)
                    continue
                spec = dict(raw_spec)
                if "name" not in spec and spec.get("sheet_name") not in (None, ""):
                    spec["name"] = spec.pop("sheet_name")
                    notes.append(f"sheets[{index}].sheet_name->name")
                if "rows" not in spec and "data" in spec:
                    spec["rows"] = spec.pop("data")
                    notes.append(f"sheets[{index}].data->rows")
                rows = spec.get("rows")
                if isinstance(rows, tuple):
                    rows = list(rows)
                    notes.append(f"sheets[{index}].rows:tuple->list")
                elif isinstance(rows, (set, frozenset)):
                    rows = sorted(rows, key=lambda item: repr(item))
                    notes.append(f"sheets[{index}].rows:set->deterministic_list")
                if isinstance(rows, list):
                    repaired_rows = []
                    for row_index, row in enumerate(rows):
                        if isinstance(row, tuple):
                            row = list(row)
                            notes.append(f"sheets[{index}].rows[{row_index}]:tuple->list")
                        elif isinstance(row, (set, frozenset)):
                            row = sorted(row, key=lambda item: repr(item))
                            notes.append(f"sheets[{index}].rows[{row_index}]:set->deterministic_list")
                        repaired_rows.append(row)
                    spec["rows"] = repaired_rows
                normalized_sheets.append(spec)
            args["sheets"] = normalized_sheets

    if name == "create_pptx":
        slides = args.get("slides")
        if isinstance(slides, dict):
            args["slides"] = [slides]
            notes.append("slides:object->list")

'''
    compat = replace_once(compat, anchor, block + anchor, "structured creation compat")
    compat_path.write_text(compat)


tools_path = Path("python/navixmind/tools/__init__.py")
tools = tools_path.read_text()
if "RASTACODER_V14_CREATION_ABI" not in tools:
    tools = replace_once(
        tools,
        '    "create_pdf": "create_pdf(output_path, content, title, image_paths)",',
        '    "create_pdf": "create_pdf(output_path, content) for text PDFs; use image_paths only for image-based PDFs; omit unused optional arguments",',
        "create_pdf prompt hint",
    )
    tools = replace_once(
        tools,
        '_LOCAL_MODEL_HIDDEN_ARGS = {\n    "read_docx": {"extract"},',
        '_LOCAL_MODEL_HIDDEN_ARGS = {\n    # RASTACODER_V14_CREATION_ABI\n    "create_pdf": {"title"},\n    "read_docx": {"extract"},',
        "local hidden args",
    )
    old = '''    elif name == "read_docx":
        projected["description"] = (
            "Read a DOCX file. Give only docx_path for an ordinary full read; "
            "the app chooses safe extraction defaults."
        )
'''
    new = '''    if name == "create_pdf":
        projected["description"] = (
            "Create a PDF. For a normal text PDF provide output_path and actual content text only. "
            "For an image-based PDF provide output_path and image_paths. Omit unused arguments completely."
        )
    elif name == "create_xlsx":
        # Give the small model the real nested executor ABI instead of a vague
        # array<object>. This prevents sheet_name/data aliases and set-like rows.
        props["sheets"] = {
            "type": "array",
            "description": "Workbook sheets. Use name and rows exactly.",
            "items": {
                "type": "object",
                "properties": {
                    "name": {"type": "string", "description": "Worksheet name"},
                    "rows": {
                        "type": "array",
                        "items": {"type": "array", "items": {}},
                        "description": "Rows in exact column order, e.g. [[\"Model\",\"Country\"],[\"Qwen\",\"China\"]]",
                    },
                },
                "required": ["name", "rows"],
            },
        }
        required = schema.setdefault("required", [])
        if "sheets" not in required:
            required.append("sheets")
        projected["description"] = "Create a populated XLSX workbook using output_path and sheets=[{name, rows:[[...]]}]."
    elif name == "create_pptx":
        props["slides"] = {
            "type": "array",
            "items": {
                "type": "object",
                "properties": {
                    "title": {"type": "string"},
                    "content": {"type": "string"},
                    "bullets": {"type": "array", "items": {"type": "string"}},
                    "notes": {"type": "string"},
                },
            },
        }
        projected["description"] = "Create a PPTX from output_path plus structured slides; omit unused slide fields."
    elif name == "read_docx":
        projected["description"] = (
            "Read a DOCX file. Give only docx_path for an ordinary full read; "
            "the app chooses safe extraction defaults."
        )
'''
    tools = replace_once(tools, old, new, "projected creation ABI")
    tools_path.write_text(tools)


ext_path = Path("python/navixmind/tools/extended_tools.py")
ext = ext_path.read_text()
if "RASTACODER_V14_XLSX_SEMANTIC_POSTCONDITION" not in ext:
    old = '''def create_xlsx(
    output_path: str,
    sheets: Optional[List[Dict[str, Any]]] = None,
) -> dict:
    """Create an XLSX workbook from structured rows."""
    from openpyxl import Workbook

    _ensure_parent(output_path)
    wb = Workbook()
    default = wb.active
    configured = sheets or [{"name": "Sheet1", "rows": []}]
    for i, spec in enumerate(configured):
        ws = default if i == 0 else wb.create_sheet()
        ws.title = str(spec.get("name") or f"Sheet{i + 1}")[:31]
        for row in spec.get("rows", []) or []:
            ws.append(list(row) if isinstance(row, (list, tuple)) else [row])
    wb.save(output_path)
    return {"success": True, "output_path": output_path, "sheet_names": wb.sheetnames}
'''
    new = '''def create_xlsx(
    output_path: str,
    sheets: Optional[List[Dict[str, Any]]] = None,
) -> dict:
    """Create an XLSX workbook and verify requested structured rows survived."""
    # RASTACODER_V14_XLSX_SEMANTIC_POSTCONDITION
    from openpyxl import Workbook, load_workbook

    _ensure_parent(output_path)
    wb = Workbook()
    default = wb.active
    configured = sheets or [{"name": "Sheet1", "rows": []}]
    if not isinstance(configured, (list, tuple)) or not configured:
        raise ToolError("create_xlsx sheets must be a non-empty list")

    expected = []
    for i, raw_spec in enumerate(configured):
        if not isinstance(raw_spec, dict):
            raise ToolError(f"create_xlsx sheet #{i + 1} must be an object")
        spec = dict(raw_spec)
        name = spec.get("name") if spec.get("name") not in (None, "") else spec.get("sheet_name")
        rows = spec.get("rows") if "rows" in spec else spec.get("data", [])
        if rows is None:
            rows = []
        if not isinstance(rows, (list, tuple)):
            raise ToolError(f"create_xlsx sheet #{i + 1} rows must be a list of rows")
        normalized_rows = []
        for row_index, row in enumerate(rows):
            if isinstance(row, (list, tuple)):
                values = list(row)
            elif isinstance(row, (set, frozenset)):
                values = sorted(row, key=lambda item: repr(item))
            else:
                values = [row]
            normalized_rows.append(values)
        ws = default if i == 0 else wb.create_sheet()
        ws.title = str(name or f"Sheet{i + 1}")[:31]
        for values in normalized_rows:
            ws.append(values)
        expected.append((ws.title, normalized_rows))

    wb.save(output_path)
    wb.close()

    # Reopen and verify dimensions plus every requested cell value. A merely
    # existing XLSX is not considered success when the requested data vanished.
    check = load_workbook(output_path, read_only=True, data_only=False)
    try:
        for sheet_name, rows in expected:
            if sheet_name not in check.sheetnames:
                raise ToolError(f"create_xlsx verification failed: missing sheet {sheet_name}")
            ws = check[sheet_name]
            for r_idx, row in enumerate(rows, start=1):
                for c_idx, wanted in enumerate(row, start=1):
                    actual = ws.cell(row=r_idx, column=c_idx).value
                    if actual != wanted:
                        raise ToolError(
                            f"create_xlsx verification failed at {sheet_name}!R{r_idx}C{c_idx}: "
                            f"expected {wanted!r}, got {actual!r}"
                        )
        sheet_names = list(check.sheetnames)
    finally:
        check.close()

    return {
        "success": True,
        "output_path": output_path,
        "sheet_names": sheet_names,
        "row_counts": {name: len(rows) for name, rows in expected},
        "verified_cells": sum(len(row) for _, rows in expected for row in rows),
        "semantic_verified": True,
    }
'''
    ext = replace_once(ext, old, new, "create_xlsx implementation")
    ext_path.write_text(ext)

print("Applied V14 structured creation and XLSX semantic contracts")
