from pathlib import Path

ROOT = Path('.')


def read(path):
    return (ROOT / path).read_text(encoding='utf-8')


def write(path, text):
    (ROOT / path).write_text(text, encoding='utf-8')


def replace_once(path, old, new):
    text = read(path)
    if new in text:
        return
    if old not in text:
        raise RuntimeError(f'V15 anchor not found in {path}: {old[:120]!r}')
    write(path, text.replace(old, new, 1))


def replace_between(path, start, end, new_block):
    text = read(path)
    if new_block in text:
        return
    a = text.find(start)
    if a < 0:
        raise RuntimeError(f'V15 start anchor not found in {path}: {start!r}')
    b = text.find(end, a)
    if b < 0:
        raise RuntimeError(f'V15 end anchor not found in {path}: {end!r}')
    write(path, text[:a] + new_block + text[b:])


# Android Python runtime: Python 3.13 enables the current Android ARM64
# curl-cffi wheel. Chaquopy 16.1 also fixes Python 3.13 native SOABI loading.
replace_once(
    'android/build.gradle',
    'classpath "com.chaquo.python:gradle:16.0.0"',
    'classpath "com.chaquo.python:gradle:16.1.0"',
)
replace_once('android/app/build.gradle', 'version "3.11"', 'version "3.13"')
replace_once(
    'android/app/build.gradle',
    '                install "yt-dlp>=2023.12.0"',
    '                install "yt-dlp>=2026.7.4"\n                install "curl-cffi==0.16.2"',
)
replace_once(
    'android/app/build.gradle',
    '                install "numpy>=1.24.0"\n                install "pandas>=2.0.0"\n                install "matplotlib==3.6.0"',
    '                install "numpy==1.26.2"\n                install "pandas==2.1.3"\n                install "matplotlib==3.8.4"',
)
replace_once(
    'android/app/build.gradle',
    'chaquopy {\n    sourceSets {',
    'chaquopy {\n    // V15: python-pptx creates notes masters from packaged XML templates.\n'
    '    // Force pptx resources onto the filesystem for notesMaster.xml access.\n'
    '    defaultConfig {\n        extractPackages("pptx")\n    }\n    sourceSets {',
)

req = read('python/requirements.txt')
req = req.replace('yt-dlp>=2023.12.0', 'yt-dlp>=2026.7.4\ncurl-cffi==0.16.2')
req = req.replace('numpy>=1.24.0', 'numpy==1.26.2')
req = req.replace('pandas>=2.0.0', 'pandas==2.1.3')
req = req.replace('matplotlib==3.6.0', 'matplotlib==3.8.4')
write('python/requirements.txt', req)

# ZIP extraction: a zero-byte archive member is a valid file. Preserve each
# archive-declared size so the common postcondition layer can allow only
# legitimately empty extract_zip outputs while retaining strict checks elsewhere.
replace_once(
    'python/navixmind/tools/extended_tools.py',
    '    extracted: List[str] = []\n    try:',
    '    extracted: List[str] = []\n    extracted_sizes: Dict[str, int] = {}\n    try:',
)
replace_once(
    'python/navixmind/tools/extended_tools.py',
    '''                extracted.append(dest)\n        return {\n            "success": True,\n            "zip_path": zip_path,\n            "output_dir": root_real,\n            "file_count": len(extracted),\n            "output_paths": extracted[:2000],\n        }''',
    '''                extracted.append(dest)\n                extracted_sizes[dest] = int(info.file_size)\n        return {\n            "success": True,\n            "zip_path": zip_path,\n            "output_dir": root_real,\n            "file_count": len(extracted),\n            "output_paths": extracted[:2000],\n            "output_sizes": {p: extracted_sizes[p] for p in extracted[:2000]},\n        }''',
)
replace_once(
    'python/navixmind/tools/__init__.py',
    '''def _verify_output_artifact(path: Any) -> int:\n    import os\n    if not isinstance(path, str) or not path or not os.path.isfile(path):\n        raise ToolError(f"[TOOL_POSTCONDITION_ERROR] Output file missing: {path}")\n    size = os.path.getsize(path)\n    if size <= 0:\n        raise ToolError(f"[TOOL_POSTCONDITION_ERROR] Output file is empty: {path}")''',
    '''def _verify_output_artifact(path: Any, allow_empty: bool = False) -> int:\n    import os\n    if not isinstance(path, str) or not path or not os.path.isfile(path):\n        raise ToolError(f"[TOOL_POSTCONDITION_ERROR] Output file missing: {path}")\n    size = os.path.getsize(path)\n    if size <= 0 and not allow_empty:\n        raise ToolError(f"[TOOL_POSTCONDITION_ERROR] Output file is empty: {path}")''',
)
replace_once(
    'python/navixmind/tools/__init__.py',
    '    sizes = {p: _verify_output_artifact(p) for p in dict.fromkeys(paths)}',
    '''    expected_sizes = result.get("output_sizes") if isinstance(result.get("output_sizes"), dict) else {}\n    sizes = {\n        p: _verify_output_artifact(\n            p,\n            allow_empty=(tool_name == "extract_zip" and expected_sizes.get(p) == 0),\n        )\n        for p in dict.fromkeys(paths)\n    }''',
)

# Image format aliases: Pillow wants JPEG rather than JPG.
replace_once(
    'python/navixmind/tools/extended_tools.py',
    '''    ext = os.path.splitext(output_path)[1].lower()\n    save_format = str(params.get("format", "")).upper() or {\n        ".jpg": "JPEG", ".jpeg": "JPEG", ".png": "PNG", ".webp": "WEBP", ".bmp": "BMP", ".gif": "GIF", ".tif": "TIFF", ".tiff": "TIFF",\n    }.get(ext, "PNG")\n    if save_format in {"JPEG", "JPG"}:''',
    '''    ext = os.path.splitext(output_path)[1].lower()\n    aliases = {\n        "jpg": "JPEG", "jpeg": "JPEG", "png": "PNG", "webp": "WEBP",\n        "bmp": "BMP", "gif": "GIF", "tif": "TIFF", "tiff": "TIFF",\n    }\n    requested_format = str(params.get("format", "")).strip().lower().lstrip(".")\n    if requested_format:\n        save_format = aliases.get(requested_format)\n        if save_format is None:\n            raise ToolError(\n                f"Unsupported image format: {params.get('format')}. "\n                "Use jpg/jpeg, png, webp, bmp, gif, tif, or tiff."\n            )\n    else:\n        save_format = aliases.get(ext.lstrip("."), "PNG")\n    if save_format == "JPEG":''',
)

# XLSX values/formulas: preserve formulas when cached values are absent.
read_xlsx_block = r'''def read_xlsx(xlsx_path: str, sheet: str = None, range: str = None, extract: str = "values") -> dict:
    """Read spreadsheet values/formulas without losing uncached formula cells."""
    from openpyxl import load_workbook

    validate_file_for_processing(xlsx_path, 'document')
    extract = str(extract or "values").lower()
    if extract not in {"values", "formulas", "all"}:
        raise ToolError("read_xlsx extract must be values, formulas, or all")

    try:
        formula_wb = load_workbook(xlsx_path, data_only=False)
        value_wb = load_workbook(xlsx_path, data_only=True) if extract in {"values", "all"} else None
        result = {
            "path": xlsx_path,
            "sheet_names": formula_wb.sheetnames,
            "sheet_count": len(formula_wb.sheetnames),
        }
        sheet_key = str(sheet) if sheet is not None else None
        if sheet_key is not None:
            if sheet_key.isdigit():
                idx = int(sheet_key)
                if idx >= len(formula_wb.sheetnames):
                    raise ToolError(f"Sheet index {idx} out of range. Available: {len(formula_wb.sheetnames)} sheets.")
                sheets_to_read = [formula_wb.sheetnames[idx]]
            elif sheet_key in formula_wb.sheetnames:
                sheets_to_read = [sheet_key]
            else:
                raise ToolError(f"Sheet '{sheet_key}' not found. Available: {formula_wb.sheetnames}")
        else:
            sheets_to_read = formula_wb.sheetnames

        max_rows = PROCESSING_LIMITS.get('xlsx_rows', 100_000)
        sheets_data = {}
        total_fallbacks = 0
        for sheet_name in sheets_to_read:
            fws = formula_wb[sheet_name]
            vws = value_wb[sheet_name] if value_wb is not None else None
            formula_cells = fws[range] if range else fws.iter_rows()
            rows_data = []
            formula_rows = []
            fallbacks = []
            row_count = 0
            for frow in formula_cells:
                if row_count >= max_rows:
                    rows_data.append(["[Truncated — max rows exceeded]"])
                    if extract == "all":
                        formula_rows.append(["[Truncated — max rows exceeded]"])
                    break
                current_values = []
                current_formulas = []
                for fcell in frow:
                    raw_formula = fcell.value
                    is_formula = fcell.data_type == 'f' or (
                        isinstance(raw_formula, str) and raw_formula.startswith('=')
                    )
                    current_formulas.append(raw_formula)
                    if vws is None:
                        current_values.append(raw_formula)
                    else:
                        cached = vws[fcell.coordinate].value
                        if is_formula and cached is None:
                            current_values.append(raw_formula)
                            fallbacks.append({
                                "cell": fcell.coordinate,
                                "formula": raw_formula,
                                "cached_value": None,
                            })
                        else:
                            current_values.append(cached)
                if extract == "formulas":
                    rows_data.append(current_formulas)
                else:
                    rows_data.append(current_values)
                    if extract == "all":
                        formula_rows.append(current_formulas)
                row_count += 1
            entry = {
                "rows": rows_data,
                "row_count": row_count,
                "dimensions": fws.dimensions,
                "has_uncached_formulas": bool(fallbacks),
            }
            if extract == "all":
                entry["formula_rows"] = formula_rows
            if fallbacks:
                entry["formula_fallbacks"] = fallbacks[:2000]
                total_fallbacks += len(fallbacks)
            sheets_data[sheet_name] = entry
        result["sheets"] = sheets_data
        result["uncached_formula_count"] = total_fallbacks
        formula_wb.close()
        if value_wb is not None:
            value_wb.close()
        return result
    except ToolError:
        raise
    except Exception as e:
        raise ToolError(f"Failed to read XLSX: {str(e)}")


'''
replace_between(
    'python/navixmind/tools/documents.py',
    'def read_xlsx(',
    'def modify_xlsx(',
    read_xlsx_block,
)

print('V15 document/runtime patch applied.')
